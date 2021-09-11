from openpyxl import load_workbook


class Parser:

    def __init__(self, file, input_user, group_names, dictionary, days, day, schedule_day, line, pair, union_1,
                 union_2):
        self.file = file
        self.input_user = input_user
        self.group_names = group_names
        self.dictionary = dictionary
        self.days = days
        self.day = day
        self.schedule_day = schedule_day
        self.line = line
        self.pair = pair
        self.union_1 = union_1
        self.union_2 = union_2

        Parser.Adding(self)

    @staticmethod
    def Error():
        exit()

    def Logic(self):
        index = self.group_names.index(self.input_user) + 3
        while self.line < 132:

            if self.pair == 5:
                self.pair = 0
                self.day = days[0]
                self.dictionary[self.day] = self.schedule_day
                self.schedule_day = '---------------\n'
                del self.days[0]

            if self.pair % 5 == 0:
                self.day = self.days[0]
                self.schedule_day = '{0}:\n---------------\n'.format(self.day.title())

            if index > 3:
                cell = self.file.cell(row=self.line, column=index)
                for mergedCell in self.file.merged_cells.ranges:
                    if cell.coordinate in mergedCell:
                        self.union_1 = 1
                cell = self.file.cell(row=self.line + 2, column=index)
                for mergedCell in self.file.merged_cells.ranges:
                    if cell.coordinate in mergedCell:
                        self.union_2 = 1
                if (self.union_1 == 1) and (self.file[self.line + 1][index].value is not None):
                    self.union_1 = 0
                if (self.union_2 == 1) and (self.file[self.line + 2][index].value is not None):
                    self.union_2 = 0

            if (self.file[self.line][index - self.union_1].value is None) and (
                    self.file[self.line + 1][index - self.union_1].value is None):
                self.schedule_day += 'Неч. Пары нет\n'

            if (self.file[self.line][index - self.union_1].value is not None) and (
                    self.file[self.line + 1][index - self.union_1].value is not None):
                self.schedule_day += 'Неч. {0} {1}\n'.format(self.file[self.line][index - self.union_1].value,
                                                             self.file[self.line + 1][index - self.union_1].value)

            if (self.file[self.line][index - self.union_1].value is None) and (
                    self.file[self.line + 1][index - self.union_1].value is not None):
                self.schedule_day += 'Всегда {0} {1}\n'.format(self.file[self.line + 1][index - self.union_1].value,
                                                               self.file[self.line + 2][index - self.union_1].value)
                self.schedule_day += '---------------\n'

            elif (self.file[self.line + 2][index - self.union_2].value is not None) and (
                    self.file[self.line + 3][index - self.union_2].value is not None):
                self.schedule_day += 'Чет. {0} {1}\n'.format(self.file[self.line + 2][index - self.union_2].value,
                                                             self.file[self.line + 3][index - self.union_2].value)
                self.schedule_day += '---------------\n'

            if (self.file[self.line + 2][index - self.union_2].value is None) and (
                    self.file[self.line + 3][index - self.union_2].value is None):
                self.schedule_day += 'Чет. Пары нет\n'
                self.schedule_day += '---------------\n'

            self.line += 4
            self.pair += 1
            self.union_1, self.union_2 = 0, 0

        self.dictionary['saturday'] = self.schedule_day

    def Adding(self):
        for a in range(3, 19):
            self.group_names.append(self.file[11][a].value)
        Parser.Examination(self)

    def Examination(self):
        if self.input_user in self.group_names:
            Parser.Logic(self)
        else:
            Parser.Error()


file = load_workbook('Schedule.xlsx').active
group_names = []
dictionary = {'monday': [],
              'tuesday': [],
              'wednesday': [],
              'thursday': [],
              'friday': [],
              'saturday': []}
days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
day = ''
input_user = input().upper()
schedule_day = '---------------\n'
line, pair, union_1, union_2 = 12, 0, 0, 0

parser = Parser(file, input_user, group_names, dictionary, days, day, schedule_day, line, pair, union_1, union_2)
